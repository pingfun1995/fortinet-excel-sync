"""
نظارت روی یک فایل اکسل و همگام‌سازی خودکار یک یا چند فایروال Fortinet از طریق SSH:

  شیت آدرس‌ها (Name/IP/Firewall):
    - ردیف‌های جدید  -> address object ساخته می‌شود
    - ردیف‌هایی که IP‌شان از قبل هست -> رد می‌شوند (بدون خطا)
    - ردیف‌های حذف‌شده که خودِ اسکریپت ساخته بود -> از فایروال هم حذف می‌شوند

  شیت تانل‌ها/Tunnels (Name/RemoteGateway/PSK/LocalSubnet/RemoteSubnet/...):
    - برای هر ردیف یک IPsec VPN Tunnel از نوع route-based کامل ساخته می‌شود:
      Phase1 + Phase2 + Static Route + دو Policy عبور ترافیک (رفت و برگشت)
      + دو Address Object برای ساب‌نت لوکال/ریموت
    - تانل‌های حذف‌شده از اکسل (که خودِ اسکریپت ساخته بود) با ترتیب درست حذف می‌شوند

  - بین هر دستور یک تاخیر (پیش‌فرض ۳ ثانیه) گذاشته می‌شود تا فایروال هنگ نکند
  - بعد از هر اجرا یک گزارش اکسل ذخیره می‌شود

اجرا (به‌صورت دستی/کنسول):
    python fortinet_address_sync.py

برای اجرای دائمی به‌صورت سرویس ویندوز از fortinet_sync_service.py استفاده کن.

قبل از اجرا:
    1. pip install -r requirements.txt
    2. کپی از config.json.example به config.json و پر کردن اطلاعات واقعی

⚠️ قبل از استفاده روی فایروال واقعی، حتماً روی یک فایروال آزمایشی یا توی
   maintenance window تست کن — این اسکریپت Policy و VPN Tunnel هم می‌سازد.
"""

from __future__ import annotations

import ipaddress
import json
import logging
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from colorama import Fore, Style, init as colorama_init
from netmiko import ConnectHandler
from netmiko.exceptions import NetmikoAuthenticationException, NetmikoTimeoutException

CONFIG_PATH = Path(__file__).with_name("config.json")
REPORT_HEADERS = ["Time", "Firewall", "Kind", "Name", "Detail", "Action", "Message"]
TUNNEL_COMMENT_PREFIX = "excelsync:"


# ---------------------------------------------------------------------------
# تنظیمات و لاگ
# ---------------------------------------------------------------------------

def resolve_path(value: str) -> str:
    """مسیرهای نسبی رو به پوشه‌ی خود پروژه (کنار config.json) وصل می‌کنه، نه به
    working directory فعلی — چون وقتی به‌صورت سرویس ویندوز اجرا میشه، CWD قابل اعتماد نیست."""
    p = Path(value)
    return str(p if p.is_absolute() else (CONFIG_PATH.parent / p))


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        print(
            f"فایل config.json پیدا نشد. اول از config.json.example یک کپی به اسم "
            f"config.json بساز و اطلاعات فایروال(ها)/اکسل رو داخلش پر کن.\n"
            f"مسیر مورد انتظار: {CONFIG_PATH}"
        )
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    if not cfg.get("firewalls"):
        print("توی config.json حداقل یک فایروال باید توی لیست 'firewalls' تعریف بشه.")
        sys.exit(1)

    cfg["excel"]["path"] = resolve_path(cfg["excel"]["path"])
    cfg["state_file"] = resolve_path(cfg.get("state_file", "sync_state.json"))
    cfg["report_file"] = resolve_path(cfg.get("report_file", "reports/sync_report.xlsx"))
    cfg["log_file"] = resolve_path(cfg.get("log_file", "fortinet_sync.log"))
    return cfg


class ColorFormatter(logging.Formatter):
    LEVEL_COLORS = {
        logging.DEBUG: Fore.CYAN,
        logging.WARNING: Fore.YELLOW,
        logging.ERROR: Fore.RED,
        logging.CRITICAL: Fore.RED + Style.BRIGHT,
    }
    KEYWORD_COLORS = [
        ("اضافه شد", Fore.GREEN),
        ("آپدیت شد", Fore.BLUE),
        ("تکمیل شد", Fore.BLUE),
        ("حذف شد", Fore.MAGENTA),
        ("رد شد", Fore.YELLOW),
        ("خطا", Fore.RED),
    ]

    def format(self, record: logging.LogRecord) -> str:
        message = super().format(record)
        color = self.LEVEL_COLORS.get(record.levelno, "")
        text = record.getMessage()
        for keyword, kw_color in self.KEYWORD_COLORS:
            if keyword in text:
                color = kw_color
                break
        return f"{color}{message}{Style.RESET_ALL}" if color else message


def setup_logging(log_file: str, colorize: bool = True) -> None:
    if colorize:
        colorama_init()

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(
        ColorFormatter("%(asctime)s [%(levelname)s] %(message)s")
        if colorize
        else logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    )

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

    root = logging.getLogger()
    root.setLevel(logging.INFO)
    root.handlers.clear()
    root.addHandler(console_handler)
    root.addHandler(file_handler)


def print_banner() -> None:
    print(f"{Fore.CYAN}{Style.BRIGHT}")
    print("=" * 52)
    print("  Fortinet Excel Sync — Addresses + IPsec Tunnels")
    print("=" * 52)
    print(Style.RESET_ALL)


def print_summary(fw_name: str, label: str, stats: dict) -> None:
    print(f"{Fore.CYAN}{'-' * 40}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}خلاصه {label} — فایروال: {fw_name}{Style.RESET_ALL}")
    print(f"  {Fore.GREEN}+ اضافه/تکمیل شد : {stats['added']}{Style.RESET_ALL}")
    print(f"  {Fore.BLUE}~ آپدیت شد        : {stats['updated']}{Style.RESET_ALL}")
    print(f"  {Fore.YELLOW}= رد شد           : {stats['skipped']}{Style.RESET_ALL}")
    print(f"  {Fore.MAGENTA}- حذف شد          : {stats['deleted']}{Style.RESET_ALL}")
    print(f"  {Fore.RED}x خطا             : {stats['failed']}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'-' * 40}{Style.RESET_ALL}")


def new_stats() -> dict:
    return {"added": 0, "updated": 0, "skipped": 0, "deleted": 0, "failed": 0}


def add_stats(target: dict, extra: dict) -> None:
    for key in target:
        target[key] += extra.get(key, 0)


# ---------------------------------------------------------------------------
# خواندن اکسل — شیت آدرس‌ها
# ---------------------------------------------------------------------------

def file_fingerprint(path: Path) -> str:
    stat = path.stat()
    return f"{stat.st_mtime_ns}:{stat.st_size}"


def read_excel_entries(cfg: dict) -> tuple[list[dict], bool]:
    excel_cfg = cfg["excel"]
    wb = openpyxl.load_workbook(excel_cfg["path"], data_only=True)
    sheet_name = excel_cfg.get("sheet_name")
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {}
    for idx, value in enumerate(header_row):
        if value:
            headers[str(value).strip().lower()] = idx

    name_col = excel_cfg.get("name_column", "Name").strip().lower()
    ip_col = excel_cfg.get("ip_column", "IP").strip().lower()
    fw_col = excel_cfg.get("firewall_column", "Firewall").strip().lower()

    if ip_col not in headers:
        raise ValueError(
            f"ستون '{excel_cfg.get('ip_column')}' توی هدر اکسل پیدا نشد. "
            f"ستون‌های موجود: {list(headers.keys())}"
        )

    ip_idx = headers[ip_col]
    name_idx = headers.get(name_col)
    fw_idx = headers.get(fw_col)

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ip_idx >= len(row):
            continue
        ip_value = row[ip_idx]
        if not ip_value:
            continue
        ip_value = str(ip_value).strip()

        name_value = ""
        if name_idx is not None and name_idx < len(row) and row[name_idx]:
            name_value = str(row[name_idx]).strip()
        if not name_value:
            name_value = f"AUTO_{ip_value.replace('/', '_').replace(':', '_')}"

        fw_value = ""
        if fw_idx is not None and fw_idx < len(row) and row[fw_idx]:
            fw_value = str(row[fw_idx]).strip()

        entries.append({"name": name_value, "ip": ip_value, "firewall": fw_value})

    return entries, fw_idx is not None


# ---------------------------------------------------------------------------
# خواندن اکسل — شیت تانل‌ها
# ---------------------------------------------------------------------------

def read_tunnel_entries(cfg: dict) -> tuple[list[dict], bool]:
    ipsec_cfg = cfg.get("ipsec")
    if not ipsec_cfg:
        return [], False

    wb = openpyxl.load_workbook(cfg["excel"]["path"], data_only=True)
    sheet_name = ipsec_cfg.get("sheet_name", "Tunnels")
    if sheet_name not in wb.sheetnames:
        return [], False
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {}
    for idx, value in enumerate(header_row):
        if value:
            headers[str(value).strip().lower()] = idx

    def col(key: str, default: str) -> str:
        return ipsec_cfg.get(key, default).strip().lower()

    name_col = col("name_column", "Name")
    gw_col = col("remote_gateway_column", "RemoteGateway")
    psk_col = col("psk_column", "PSK")
    local_col = col("local_subnet_column", "LocalSubnet")
    remote_col = col("remote_subnet_column", "RemoteSubnet")
    if_col = col("interface_column", "Interface")
    fw_col = col("firewall_column", "Firewall")

    required = {"name": name_col, "gw": gw_col, "psk": psk_col, "local": local_col, "remote": remote_col}
    missing = [v for v in required.values() if v not in headers]
    if missing:
        raise ValueError(f"ستون‌های لازم توی شیت '{sheet_name}' پیدا نشدن: {missing}")

    idxs = {k: headers[v] for k, v in required.items()}
    if_idx = headers.get(if_col)
    fw_idx = headers.get(fw_col)

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idxs["name"] >= len(row) or not row[idxs["name"]]:
            continue
        name = str(row[idxs["name"]]).strip()
        if not name:
            continue

        gw = str(row[idxs["gw"]]).strip() if idxs["gw"] < len(row) and row[idxs["gw"]] else ""
        psk = str(row[idxs["psk"]]).strip() if idxs["psk"] < len(row) and row[idxs["psk"]] else ""
        local_subnet = str(row[idxs["local"]]).strip() if idxs["local"] < len(row) and row[idxs["local"]] else ""
        remote_subnet = str(row[idxs["remote"]]).strip() if idxs["remote"] < len(row) and row[idxs["remote"]] else ""

        if not (gw and psk and local_subnet and remote_subnet):
            logging.warning(f"ردیف تانل ناقص رد شد (RemoteGateway/PSK/LocalSubnet/RemoteSubnet لازمه): {name}")
            continue

        interface = ""
        if if_idx is not None and if_idx < len(row) and row[if_idx]:
            interface = str(row[if_idx]).strip()
        if not interface:
            interface = ipsec_cfg.get("default_wan_interface", "wan1")

        fw_value = ""
        if fw_idx is not None and fw_idx < len(row) and row[fw_idx]:
            fw_value = str(row[fw_idx]).strip()

        entries.append({
            "name": name,
            "remote_gw": gw,
            "psk": psk,
            "local_subnet": local_subnet,
            "remote_subnet": remote_subnet,
            "interface": interface,
            "firewall": fw_value,
        })

    return entries, fw_idx is not None


# ---------------------------------------------------------------------------
# ابزارهای مشترک (فیلتر فایروال، اعتبارسنجی IP)
# ---------------------------------------------------------------------------

def entries_for_firewall(entries: list[dict], fw_name: str, has_firewall_column: bool) -> list[dict]:
    if not has_firewall_column:
        return entries
    fw_name_lower = fw_name.strip().lower()
    result = []
    for entry in entries:
        fw_value = entry["firewall"].strip().lower()
        if not fw_value or fw_value in ("all", "*") or fw_value == fw_name_lower:
            result.append(entry)
    return result


def warn_unknown_firewalls(entries: list[dict], fw_names: list[str], has_firewall_column: bool, kind: str) -> None:
    if not has_firewall_column:
        return
    known = {n.strip().lower() for n in fw_names} | {"", "all", "*"}
    unknown = {entry["firewall"] for entry in entries if entry["firewall"].strip().lower() not in known}
    for value in unknown:
        logging.warning(f"[{kind}] مقدار ستون Firewall ناشناخته و رد شد: '{value}'")


def normalize_subnet(ip_value: str) -> tuple[str, str]:
    ip_value = ip_value.strip()
    if "/" in ip_value:
        network = ipaddress.ip_network(ip_value, strict=False)
        return str(network.network_address), str(network.netmask)
    ipaddress.ip_address(ip_value)  # اعتبارسنجی
    return ip_value, "255.255.255.255"


# ---------------------------------------------------------------------------
# ارتباط با فایروال — Address Objects
# ---------------------------------------------------------------------------

def fetch_existing_objects(conn) -> dict[str, tuple[str, str]]:
    output = conn.send_command("show firewall address", read_timeout=30)
    objects: dict[str, tuple[str, str]] = {}
    current_name = None
    for raw_line in output.splitlines():
        line = raw_line.strip()
        if line.startswith('edit "'):
            current_name = line[len('edit "'):].rstrip('"')
        elif line.startswith("set subnet") and current_name:
            parts = line.split()
            if len(parts) >= 4:
                objects[current_name] = (parts[2], parts[3])
            elif len(parts) == 3 and "/" in parts[2]:
                net = ipaddress.ip_network(parts[2], strict=False)
                objects[current_name] = (str(net.network_address), str(net.netmask))
        elif line == "next":
            current_name = None
    return objects


def create_address_object(conn, name: str, network: str, netmask: str) -> None:
    commands = [
        "config firewall address",
        f'edit "{name}"',
        f"set subnet {network} {netmask}",
        "next",
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


def delete_address_object(conn, name: str) -> None:
    commands = [
        "config firewall address",
        f'delete "{name}"',
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


# ---------------------------------------------------------------------------
# ارتباط با فایروال — IPsec Tunnels (route-based)
# ---------------------------------------------------------------------------

def fetch_existing_names(conn, show_command: str) -> set[str]:
    output = conn.send_command(show_command, read_timeout=30)
    names = set()
    for raw_line in output.splitlines():
        line = raw_line.strip()
        if line.startswith('edit "'):
            names.add(line[len('edit "'):].rstrip('"'))
    return names


def fetch_existing_routes_by_comment(conn) -> dict[str, str]:
    """برمی‌گردونه: {tunnel_name: route_id} فقط برای route‌هایی که این اسکریپت ساخته."""
    output = conn.send_command("show router static", read_timeout=30)
    result: dict[str, str] = {}
    current_id = None
    current_comment = None
    for raw_line in output.splitlines():
        line = raw_line.strip()
        if line.startswith("edit "):
            parts = line.split()
            current_id = parts[1] if len(parts) > 1 else None
            current_comment = None
        elif line.startswith("set comment"):
            parts = line.split(None, 2)
            if len(parts) == 3:
                current_comment = parts[2].strip('"')
        elif line == "next":
            if current_id and current_comment and current_comment.startswith(TUNNEL_COMMENT_PREFIX):
                tunnel_name = current_comment[len(TUNNEL_COMMENT_PREFIX):]
                result[tunnel_name] = current_id
            current_id = None
            current_comment = None
    return result


def fetch_existing_policy_ids_by_name(conn) -> dict[str, str]:
    output = conn.send_command("show firewall policy", read_timeout=30)
    result: dict[str, str] = {}
    current_id = None
    for raw_line in output.splitlines():
        line = raw_line.strip()
        if line.startswith("edit "):
            parts = line.split()
            current_id = parts[1] if len(parts) > 1 else None
        elif line.startswith("set name"):
            parts = line.split(None, 2)
            if len(parts) == 3 and current_id:
                result[parts[2].strip('"')] = current_id
        elif line == "next":
            current_id = None
    return result


def create_phase1(conn, name: str, ipsec_cfg: dict, remote_gw: str, psk: str, wan_interface: str) -> None:
    commands = [
        "config vpn ipsec phase1-interface",
        f'edit "{name}"',
        f'set interface "{wan_interface}"',
        f"set ike-version {ipsec_cfg.get('ike_version', '2')}",
        "set peertype any",
        "set net-device disable",
        f"set proposal {ipsec_cfg.get('phase1_proposal', 'aes256-sha256')}",
        f"set dhgrp {ipsec_cfg.get('dhgrp', '14')}",
        f"set remote-gw {remote_gw}",
        f'set psksecret "{psk}"',
        "next",
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


def delete_phase1(conn, name: str) -> None:
    conn.send_config_set(["config vpn ipsec phase1-interface", f'delete "{name}"', "end"], read_timeout=30)


def create_phase2(conn, name: str, ipsec_cfg: dict, local_net: str, local_mask: str, remote_net: str, remote_mask: str) -> None:
    commands = [
        "config vpn ipsec phase2-interface",
        f'edit "{name}"',
        f'set phase1name "{name}"',
        f"set proposal {ipsec_cfg.get('phase2_proposal', 'aes256-sha256')}",
        f"set src-subnet {local_net} {local_mask}",
        f"set dst-subnet {remote_net} {remote_mask}",
        "next",
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


def delete_phase2(conn, name: str) -> None:
    conn.send_config_set(["config vpn ipsec phase2-interface", f'delete "{name}"', "end"], read_timeout=30)


def create_static_route(conn, dst_net: str, dst_mask: str, device: str, comment: str) -> None:
    commands = [
        "config router static",
        "edit 0",
        f"set dst {dst_net} {dst_mask}",
        f'set device "{device}"',
        f'set comment "{comment}"',
        "next",
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


def delete_static_route(conn, route_id: str) -> None:
    conn.send_config_set(["config router static", f"delete {route_id}", "end"], read_timeout=30)


def create_policy(conn, name: str, srcintf: str, dstintf: str, srcaddr: str, dstaddr: str, service: str = "ALL") -> None:
    commands = [
        "config firewall policy",
        "edit 0",
        f'set name "{name}"',
        f'set srcintf "{srcintf}"',
        f'set dstintf "{dstintf}"',
        f'set srcaddr "{srcaddr}"',
        f'set dstaddr "{dstaddr}"',
        "set action accept",
        'set schedule "always"',
        f'set service "{service}"',
        "next",
        "end",
    ]
    conn.send_config_set(commands, read_timeout=30)


def delete_policy(conn, policy_id: str) -> None:
    conn.send_config_set(["config firewall policy", f"delete {policy_id}", "end"], read_timeout=30)


# ---------------------------------------------------------------------------
# state (اینکه کدوم آبجکت‌ها رو خود اسکریپت ساخته)
# ---------------------------------------------------------------------------

def load_state(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_state(path: Path, state: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def normalize_fw_state(raw: dict | None) -> dict:
    """پشتیبانی از فرمت قدیمی state (قبل از اضافه شدن تانل‌ها) که مستقیم name->subnet بود."""
    if not raw:
        return {"addresses": {}, "tunnels": {}}
    if "addresses" in raw or "tunnels" in raw:
        return {"addresses": raw.get("addresses", {}), "tunnels": raw.get("tunnels", {})}
    return {"addresses": raw, "tunnels": {}}


# ---------------------------------------------------------------------------
# گزارش اکسل
# ---------------------------------------------------------------------------

def append_report(report_path: Path, rows: list[dict]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    if report_path.exists():
        wb = openpyxl.load_workbook(report_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SyncReport"
        ws.append(REPORT_HEADERS)

    for row in rows:
        ws.append([row.get(h, "") for h in REPORT_HEADERS])

    wb.save(report_path)


def report_row(now_str: str, fw_name: str, kind: str, name: str, detail: str, action: str, message: str) -> dict:
    return {"Time": now_str, "Firewall": fw_name, "Kind": kind, "Name": name,
            "Detail": detail, "Action": action, "Message": message}


# ---------------------------------------------------------------------------
# منطق sync — Address Objects (روی یک کانکشن باز)
# ---------------------------------------------------------------------------

def sync_addresses(conn, fw_name: str, entries: list[dict], managed_prev: dict, report_rows: list[dict],
                    delay: float, delete_removed: bool) -> tuple[dict, dict]:
    stats = new_stats()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    managed_now: dict[str, list[str]] = {}

    existing = fetch_existing_objects(conn)
    existing_subnets = set(existing.values())
    logging.info(f"[{fw_name}] {len(existing)} آبجکت آدرس موجود روی فایروال.")

    desired: dict[str, tuple[str, str]] = {}
    ip_display: dict[str, str] = {}
    for entry in entries:
        try:
            network, netmask = normalize_subnet(entry["ip"])
        except ValueError as exc:
            logging.warning(f"[{fw_name}] IP نامعتبر رد شد ({entry['ip']}): {exc}")
            stats["failed"] += 1
            report_rows.append(report_row(now_str, fw_name, "Address", entry["name"], entry["ip"], "InvalidIP", str(exc)))
            continue
        desired[entry["name"]] = (network, netmask)
        ip_display[entry["name"]] = entry["ip"]

    for name, subnet in desired.items():
        network, netmask = subnet
        ip_value = ip_display[name]
        current = existing.get(name)

        if current == subnet:
            managed_now[name] = list(subnet)
            stats["skipped"] += 1
            logging.info(f"[{fw_name}] رد شد (از قبل درست روی فایروال هست): {name} -> {ip_value}")
            report_rows.append(report_row(now_str, fw_name, "Address", name, ip_value, "Skipped", "از قبل موجود بود"))
            continue

        if current is None and subnet in existing_subnets:
            stats["skipped"] += 1
            logging.warning(f"[{fw_name}] رد شد (این IP از قبل با اسم دیگه‌ای ثبت شده): {name} -> {ip_value}")
            report_rows.append(report_row(now_str, fw_name, "Address", name, ip_value, "Skipped", "IP با اسم دیگه‌ای موجود بود"))
            continue

        try:
            create_address_object(conn, name, network, netmask)
            existing_subnets.add(subnet)
            managed_now[name] = list(subnet)
            if current is None:
                stats["added"] += 1
                logging.info(f"[{fw_name}] اضافه شد: {name} -> {ip_value}")
                report_rows.append(report_row(now_str, fw_name, "Address", name, ip_value, "Added", ""))
            else:
                stats["updated"] += 1
                logging.info(f"[{fw_name}] آپدیت شد: {name} -> {ip_value}")
                report_rows.append(report_row(now_str, fw_name, "Address", name, ip_value, "Updated", ""))
        except Exception as exc:
            stats["failed"] += 1
            logging.error(f"[{fw_name}] خطا هنگام ثبت {name} ({ip_value}): {exc}")
            report_rows.append(report_row(now_str, fw_name, "Address", name, ip_value, "Failed", str(exc)))

        time.sleep(delay)  # فاصله بین دستورات تا فایروال هنگ نکنه

    if delete_removed:
        to_delete = [n for n in managed_prev if n not in desired]
        for name in to_delete:
            if name not in existing:
                continue
            try:
                delete_address_object(conn, name)
                stats["deleted"] += 1
                logging.info(f"[{fw_name}] حذف شد (از اکسل پاک شده بود): {name}")
                report_rows.append(report_row(now_str, fw_name, "Address", name, "", "Deleted", "از اکسل حذف شده بود"))
            except Exception as exc:
                stats["failed"] += 1
                logging.error(f"[{fw_name}] خطا هنگام حذف {name} (احتمالاً داخل یه پالیسی/گروه استفاده شده): {exc}")
                managed_now[name] = managed_prev[name]
                report_rows.append(report_row(now_str, fw_name, "Address", name, "", "DeleteFailed", str(exc)))
            time.sleep(delay)
    else:
        for name, subnet in managed_prev.items():
            if name not in desired:
                managed_now.setdefault(name, subnet)

    return stats, managed_now


# ---------------------------------------------------------------------------
# منطق sync — IPsec Tunnels (روی یک کانکشن باز)
# ---------------------------------------------------------------------------

def sync_tunnels(conn, fw_name: str, entries: list[dict], managed_prev: dict, report_rows: list[dict],
                  delay: float, delete_removed: bool, ipsec_cfg: dict, lan_interface: str) -> tuple[dict, dict]:
    stats = new_stats()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    managed_now: dict[str, dict] = {}

    existing_phase1 = fetch_existing_names(conn, "show vpn ipsec phase1-interface")
    existing_phase2 = fetch_existing_names(conn, "show vpn ipsec phase2-interface")
    existing_routes = fetch_existing_routes_by_comment(conn)
    existing_policies = fetch_existing_policy_ids_by_name(conn)
    logging.info(f"[{fw_name}] {len(existing_phase1)} تانل IPsec موجود روی فایروال.")

    desired_names = set()

    for entry in entries:
        name = entry["name"]
        desired_names.add(name)

        try:
            local_net, local_mask = normalize_subnet(entry["local_subnet"])
            remote_net, remote_mask = normalize_subnet(entry["remote_subnet"])
        except ValueError as exc:
            stats["failed"] += 1
            logging.warning(f"[{fw_name}] ساب‌نت نامعتبر توی تانل {name}: {exc}")
            report_rows.append(report_row(now_str, fw_name, "Tunnel", name, "", "InvalidSubnet", str(exc)))
            continue

        local_addr_name = f"{name}_local"
        remote_addr_name = f"{name}_remote"
        policy_out_name = f"{name}_out"
        policy_in_name = f"{name}_in"

        fully_exists = (
            name in existing_phase1 and name in existing_phase2 and name in existing_routes
            and policy_out_name in existing_policies and policy_in_name in existing_policies
        )
        if fully_exists and name in managed_prev:
            stats["skipped"] += 1
            logging.info(f"[{fw_name}] تانل رد شد (از قبل کامل بود): {name}")
            report_rows.append(report_row(now_str, fw_name, "Tunnel", name, entry["remote_gw"], "Skipped", "از قبل کامل موجود بود"))
            managed_now[name] = managed_prev[name]
            continue

        try:
            existing_addrs = fetch_existing_objects(conn)

            if local_addr_name not in existing_addrs:
                create_address_object(conn, local_addr_name, local_net, local_mask)
                time.sleep(delay)
            if remote_addr_name not in existing_addrs:
                create_address_object(conn, remote_addr_name, remote_net, remote_mask)
                time.sleep(delay)

            if name not in existing_phase1:
                create_phase1(conn, name, ipsec_cfg, entry["remote_gw"], entry["psk"], entry["interface"])
                existing_phase1.add(name)
                time.sleep(delay)

            if name not in existing_phase2:
                create_phase2(conn, name, ipsec_cfg, local_net, local_mask, remote_net, remote_mask)
                existing_phase2.add(name)
                time.sleep(delay)

            if name not in existing_routes:
                create_static_route(conn, remote_net, remote_mask, name, TUNNEL_COMMENT_PREFIX + name)
                time.sleep(delay)
                existing_routes = fetch_existing_routes_by_comment(conn)

            if policy_out_name not in existing_policies:
                create_policy(conn, policy_out_name, lan_interface, name, local_addr_name, remote_addr_name)
                time.sleep(delay)
                existing_policies = fetch_existing_policy_ids_by_name(conn)

            if policy_in_name not in existing_policies:
                create_policy(conn, policy_in_name, name, lan_interface, remote_addr_name, local_addr_name)
                time.sleep(delay)
                existing_policies = fetch_existing_policy_ids_by_name(conn)

            managed_now[name] = {
                "local_addr": local_addr_name,
                "remote_addr": remote_addr_name,
                "policy_out": policy_out_name,
                "policy_in": policy_in_name,
            }
            stats["added"] += 1
            logging.info(f"[{fw_name}] تانل اضافه/تکمیل شد: {name} -> {entry['remote_gw']}")
            report_rows.append(report_row(now_str, fw_name, "Tunnel", name, entry["remote_gw"], "Added", ""))
        except Exception as exc:
            stats["failed"] += 1
            logging.error(f"[{fw_name}] خطا هنگام ساخت تانل {name}: {exc}")
            report_rows.append(report_row(now_str, fw_name, "Tunnel", name, entry.get("remote_gw", ""), "Failed", str(exc)))

    if delete_removed:
        to_delete = [n for n in managed_prev if n not in desired_names]
        for name in to_delete:
            try:
                policy_out_name = f"{name}_out"
                policy_in_name = f"{name}_in"

                existing_policies = fetch_existing_policy_ids_by_name(conn)
                if policy_out_name in existing_policies:
                    delete_policy(conn, existing_policies[policy_out_name])
                    time.sleep(delay)
                if policy_in_name in existing_policies:
                    delete_policy(conn, existing_policies[policy_in_name])
                    time.sleep(delay)

                existing_routes = fetch_existing_routes_by_comment(conn)
                if name in existing_routes:
                    delete_static_route(conn, existing_routes[name])
                    time.sleep(delay)

                if name in fetch_existing_names(conn, "show vpn ipsec phase2-interface"):
                    delete_phase2(conn, name)
                    time.sleep(delay)
                if name in fetch_existing_names(conn, "show vpn ipsec phase1-interface"):
                    delete_phase1(conn, name)
                    time.sleep(delay)

                for addr_name in (f"{name}_local", f"{name}_remote"):
                    try:
                        delete_address_object(conn, addr_name)
                    except Exception:
                        pass  # ممکنه جای دیگه‌ای هم استفاده شده باشه، مشکلی نیست
                    time.sleep(delay)

                stats["deleted"] += 1
                logging.info(f"[{fw_name}] تانل حذف شد (از اکسل پاک شده بود): {name}")
                report_rows.append(report_row(now_str, fw_name, "Tunnel", name, "", "Deleted", "از اکسل حذف شده بود"))
            except Exception as exc:
                stats["failed"] += 1
                managed_now[name] = managed_prev[name]
                logging.error(f"[{fw_name}] خطا هنگام حذف تانل {name}: {exc}")
                report_rows.append(report_row(now_str, fw_name, "Tunnel", name, "", "DeleteFailed", str(exc)))
    else:
        for name, info in managed_prev.items():
            if name not in desired_names:
                managed_now.setdefault(name, info)

    return stats, managed_now


# ---------------------------------------------------------------------------
# اجرای هر دو sync (آدرس + تانل) روی یک فایروال با یک اتصال SSH مشترک
# ---------------------------------------------------------------------------

def process_firewall(fw_cfg: dict, addr_entries: list[dict], tunnel_entries: list[dict], state: dict,
                      report_rows: list[dict], delay: float, delete_removed: bool, ipsec_cfg: dict | None,
                      lan_interface: str) -> None:
    fw_name = fw_cfg.get("name", fw_cfg["host"])
    logging.info(f"── فایروال: {fw_name} ({fw_cfg['host']}) ──")

    if not addr_entries and not tunnel_entries:
        logging.info(f"[{fw_name}] هیچ ردیفی برای این فایروال نیست، رد شد.")
        return

    device = {
        "device_type": fw_cfg.get("device_type", "fortinet"),
        "host": fw_cfg["host"],
        "username": fw_cfg["username"],
        "password": fw_cfg["password"],
        "port": fw_cfg.get("port", 22),
    }

    try:
        conn = ConnectHandler(**device)
    except (NetmikoTimeoutException, NetmikoAuthenticationException) as exc:
        logging.error(f"[{fw_name}] اتصال SSH ناموفق: {exc}")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report_rows.append(report_row(now_str, fw_name, "-", "", "", "ConnectFailed", str(exc)))
        return

    prev = normalize_fw_state(state.get(fw_name))
    managed_now = {"addresses": prev["addresses"], "tunnels": prev["tunnels"]}

    try:
        if addr_entries or prev["addresses"]:
            addr_stats, addr_managed = sync_addresses(
                conn, fw_name, addr_entries, prev["addresses"], report_rows, delay, delete_removed
            )
            managed_now["addresses"] = addr_managed
            print_summary(fw_name, "Address Objects", addr_stats)

        if tunnel_entries or prev["tunnels"]:
            if ipsec_cfg is None:
                logging.warning(f"[{fw_name}] شیت تانل پیدا شد ولی بخش 'ipsec' توی config.json تعریف نشده — رد شد.")
            else:
                tunnel_stats, tunnel_managed = sync_tunnels(
                    conn, fw_name, tunnel_entries, prev["tunnels"], report_rows, delay, delete_removed,
                    ipsec_cfg, lan_interface
                )
                managed_now["tunnels"] = tunnel_managed
                print_summary(fw_name, "IPsec Tunnels", tunnel_stats)
    finally:
        conn.disconnect()

    state[fw_name] = managed_now


# ---------------------------------------------------------------------------
# چرخه اصلی
# ---------------------------------------------------------------------------

def sync_once(cfg: dict) -> None:
    delay = cfg["sync"].get("delay_between_commands_seconds", 3)
    delete_removed = cfg["sync"].get("delete_removed_objects", True)
    state_path = Path(cfg["state_file"])
    report_path = Path(cfg["report_file"])
    ipsec_cfg = cfg.get("ipsec")
    lan_interface = (ipsec_cfg or {}).get("default_lan_interface", "internal")

    try:
        addr_entries, addr_has_fw_col = read_excel_entries(cfg)
    except Exception as exc:
        logging.error(f"خواندن شیت آدرس‌ها ناموفق بود: {exc}")
        addr_entries, addr_has_fw_col = [], False

    try:
        tunnel_entries, tunnel_has_fw_col = read_tunnel_entries(cfg)
    except Exception as exc:
        logging.error(f"خواندن شیت تانل‌ها ناموفق بود: {exc}")
        tunnel_entries, tunnel_has_fw_col = [], False

    if not addr_entries and not tunnel_entries:
        logging.info("هیچ ردیفی توی اکسل پیدا نشد.")
        return

    firewalls = cfg["firewalls"]
    fw_names = [fw.get("name", fw["host"]) for fw in firewalls]
    warn_unknown_firewalls(addr_entries, fw_names, addr_has_fw_col, "Address")
    warn_unknown_firewalls(tunnel_entries, fw_names, tunnel_has_fw_col, "Tunnel")

    state = load_state(state_path)
    report_rows: list[dict] = []

    for fw_cfg in firewalls:
        fw_name = fw_cfg.get("name", fw_cfg["host"])
        fw_addr_entries = entries_for_firewall(addr_entries, fw_name, addr_has_fw_col)
        fw_tunnel_entries = entries_for_firewall(tunnel_entries, fw_name, tunnel_has_fw_col)
        process_firewall(fw_cfg, fw_addr_entries, fw_tunnel_entries, state, report_rows, delay,
                          delete_removed, ipsec_cfg, lan_interface)

    save_state(state_path, state)
    if report_rows:
        append_report(report_path, report_rows)
        logging.info(f"گزارش اکسل ذخیره/آپدیت شد: {report_path}")


def sleep_or_stop(seconds: float, stop_event=None) -> bool:
    """می‌خوابه، مگر اینکه stop_event ست بشه. برمی‌گردونه True یعنی باید متوقف بشه."""
    if stop_event is None:
        time.sleep(seconds)
        return False
    import win32event
    result = win32event.WaitForSingleObject(stop_event, int(seconds * 1000))
    return result == win32event.WAIT_OBJECT_0


def watch(cfg: dict, stop_event=None) -> None:
    excel_path = Path(cfg["excel"]["path"])
    poll_interval = cfg["sync"].get("poll_interval_seconds", 5)

    if not excel_path.exists():
        logging.error(f"فایل اکسل پیدا نشد: {excel_path}")
        if stop_event is None:
            sys.exit(1)
        return

    logging.info(f"شروع نظارت روی {excel_path} (هر {poll_interval} ثانیه بررسی می‌شود)")
    last_fingerprint = None

    while True:
        try:
            current_fingerprint = file_fingerprint(excel_path)
            if current_fingerprint != last_fingerprint:
                if last_fingerprint is None:
                    logging.info("اجرای اولیه — همگام‌سازی با وضعیت فعلی اکسل ...")
                else:
                    logging.info("تغییر در فایل اکسل شناسایی شد — شروع sync ...")
                    sleep_or_stop(1, stop_event)  # مطمئن بشیم اکسل کامل ذخیره شده
                sync_once(cfg)
                last_fingerprint = file_fingerprint(excel_path)
        except Exception as exc:
            logging.error(f"خطای غیرمنتظره: {exc}")

        if sleep_or_stop(poll_interval, stop_event):
            logging.info("حلقه نظارت متوقف شد.")
            break


def main() -> None:
    cfg = load_config()
    setup_logging(cfg.get("log_file", "fortinet_sync.log"))
    print_banner()
    try:
        watch(cfg)
    except KeyboardInterrupt:
        logging.info("متوقف شد توسط کاربر.")


if __name__ == "__main__":
    main()
